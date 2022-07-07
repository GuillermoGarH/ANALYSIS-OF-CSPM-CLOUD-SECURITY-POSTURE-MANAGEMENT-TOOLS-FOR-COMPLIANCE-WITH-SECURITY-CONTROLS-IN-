#####################REQUIREMENTS#######################
###########pip install openpyxl
########################################################

from time import sleep

from tkinter import *
from tkinter import filedialog
from tkinter import ttk
      
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from itertools import count
from _overlapped import NULL
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Series, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import RichTextProperties

import os

cspmPath = ""
cisPath = ""
bmPath = ""
names = []
subscription = ""    

# cspmNewName = ''
# subscriptionNewName = ''
# CSPMaddOptions =  ''

posibleOptions = ''
subscriptionName = ''

def newCSPMWin():
    newCSPMwin = Toplevel(window)
    newCSPMwin.title("Nueva suscripcion")
    newCSPMwin.geometry("450x130")

    global subscriptionName
    global posibleOptions

    
    # cspmNameTag = Entry(newCSPMwin, width = 30).place(x = 110,y = 60)  
    # cspmName = Label(newCSPMwin, text = "Nombre CSPM").place(x = 40,y = 60)  
    subscriptionNameTag = Label(newCSPMwin,  text = "Elegir opcion")
    subscriptionNameTag.grid(row = 0, column = 0, padx=10, pady=10, sticky="nsew") 
    posibleOptions = ttk.Combobox(newCSPMwin,state="readonly",values=["Agregar una suscripcion a CSPM existente", "Crear nuevo CSPM desde cero"])
    posibleOptions.grid(row = 0, column = 1, padx=10, pady=10, sticky="nsew") 
    
    subscriptionNameTag = Label(newCSPMwin,  text = "Nombre Suscripcion")   
    subscriptionNameTag.grid(row = 1, column = 0, padx=10, pady=10, sticky="nsew")    
    subscriptionName = Entry(newCSPMwin,width = 40)
    subscriptionName.grid(row = 1, column = 1, padx=10, pady=10, sticky="nsew")  
    
    guardar = Button(newCSPMwin,text = "Guardar", command=newCSPMFile)
    guardar.grid(row = 2, column = 0, padx=10, pady=10, sticky="nsew")
    
def newCSPMFile():

    global subscriptionName
    global posibleOptions
    global cspmPath
        
    CSPMaddOptions = posibleOptions.get()
    subscriptionNewName = subscriptionName.get()
    
    print("option" + CSPMaddOptions)
    if CSPMaddOptions == "Crear nuevo CSPM desde cero":    
        cspm = filedialog.asksaveasfile(initialdir = "C:\\Users", title = "Guardar CSPM", filetypes = (("Archivo Excel", "*.xlsx*"), ("all files", "*.*")), defaultextension = '.xlsx')
        cspmPath = os.path.abspath(cspm.name)  
        wbNew = openpyxl.Workbook()
        wbNew.save(cspmPath)
        
    else:
        cspm = filedialog.askopenfilename(initialdir = "C:\\Users" ,title = "Seleccionar un archivo", filetypes = (("Archivo Excel", "*.xlsx*"), ("all files", "*.*")))     
        cspmPath = os.path.abspath(cspm) 

    wb = openpyxl.load_workbook(cspmPath)   
    if CSPMaddOptions == "Crear nuevo CSPM desde cero":
        del wb[wb.sheetnames[0]] 
    ws = wb.create_sheet(title=subscriptionNewName)        
    ws = wb[subscriptionNewName]
    
    columnTitles = ['Account Name', 'Standard', 'Category', 'Team in charge', 'State', 'Comments Meeting', 'Situation', 'Rule Title', 'complianceControlName', 'Risk Level', 'Message', 'Resource', 'Link to portal']
    idx = 1
    for columnTitle in columnTitles:
        ws.cell(row = 1, column = idx).value = columnTitle
        idx += 1 
    
    global names
    global subscription
    posibleNames = wb.sheetnames
    columnTitles = ['Account Name', 'Standard', 'Category', 'Team in charge', 'State', 'Comments Meeting', 'Situation', 'Rule Title', 'complianceControlName', 'Risk Level', 'Message', 'Resource', 'Link to portal']
    isCSPM = True
    names = []
    for posibleName in posibleNames:
        ws = wb[posibleName]
        isCSPM = True
        for idxColTitle in range (0, len(columnTitles) - 1):
            if ws.cell(row = 1, column = idxColTitle + 1).value != columnTitles[idxColTitle]:
                isCSPM = False
        if isCSPM == True:
            names.append(posibleName)

    posibleSubscriptions.configure(values = names)   
    
 
    wb.save(cspmPath)
    if CSPMaddOptions == "Crear nuevo CSPM desde cero":    
        label_file_explorer.configure(text="CSPM inicial creado: " + cspmPath) 
    else:
        label_file_explorer.configure(text="CSPM " + cspmPath +  "actualizado con subscripcion: " + subscriptionNewName) 

def browseFilesCSPM(): 
    cspm = filedialog.askopenfilename(initialdir = "C:\\Users" ,title = "Seleccionar un archivo", filetypes = (("Archivo Excel", "*.xlsx*"), ("all files", "*.*")))     
    global cspmPath
    cspmPath = os.path.abspath(cspm)
    wbCSPM = openpyxl.load_workbook(os.path.abspath(cspm))

    global names
    global subscription
    posibleNames = wbCSPM.sheetnames
    columnTitles = ['Account Name', 'Standard', 'Category', 'Team in charge', 'State', 'Comments Meeting', 'Situation', 'Rule Title', 'complianceControlName', 'Risk Level', 'Message', 'Resource', 'Link to portal']
    isCSPM = True
    names = []
    for posibleName in posibleNames:
        wsCSPM = wbCSPM[posibleName]
        isCSPM = True
        for idxColTitle in range (0, len(columnTitles) - 1):
            if wsCSPM.cell(row = 1, column = idxColTitle + 1).value != columnTitles[idxColTitle]:
                isCSPM = False
        if isCSPM == True:
            names.append(posibleName)

    posibleSubscriptions.configure(values = names)   
    label_file_explorer.configure(text="File Opened: "+ cspm) 

def browseFilesCIS(): 
    
    cellStandard = ''
    
    while(cellStandard != 'Azure CIS 1.3.0'):
        cis = filedialog.askopenfilename(initialdir = "C:\\Users" ,title = "Seleccionar un archivo", filetypes = (("Archivo CSV", "*.csv*"), ("all files", "*.*")))     
    
        absolutepath = os.path.abspath(cis)
        directorypath = os.path.dirname(absolutepath)
        
        wb = openpyxl.Workbook()
        del wb[wb.sheetnames[0]]
         
        with open(cis, newline='') as f_input:
            ws = wb.create_sheet(title="CIS")
            
            for row in csv.reader(f_input, delimiter=','):
                ws.append(row)
                
        wb.save(directorypath + '\\CIS.xlsx')
          
        label_file_explorer.configure(text="Archivo abierto: "+ cis, fg = "#38EB5C" )

        checkCIS = openpyxl.load_workbook(directorypath + '\\CIS.xlsx')
        sheetCIS = checkCIS.active
        cellStandard = sheetCIS.cell(row = 2, column = 2)
        cellStandard = cellStandard.value
        
        if(cellStandard != 'Azure CIS 1.3.0'):
            label_file_explorer.configure(text='Selecciona el archivo descargado con los findings del Azure CIS 1.3.0', fg = "#f00")
            
            
    #Formatear
    cleanCIS = openpyxl.load_workbook(directorypath + '\\CIS.xlsx')
    sheetCIS = cleanCIS.active
    countRow = 1
    countColumn = 1
    referenceUnhealthyColumn = 1
    referenceSeverityColumn = 1
    
    while countColumn <= sheetCIS.max_column:
        if sheetCIS.cell(row = 1, column = countColumn).value == 'resourceState':
            referenceUnhealthyColumn = countColumn
        countColumn += 1
    
    countColumn = 1        
    while countColumn <= sheetCIS.max_column:
        if sheetCIS.cell(row = 1, column = countColumn).value == 'severity':
            referenceSeverityColumn = countColumn
        countColumn += 1
                   
    #Conserva no saneadas
    while countRow <= sheetCIS.max_row:
        unhealthy = sheetCIS.cell(row = countRow, column = referenceUnhealthyColumn)
        if unhealthy.value not in ['resourceState','unhealthy']:
            sheetCIS.delete_rows(countRow)
        else:
            countRow += 1
   
    countRow = 1
    opcionSoloHigh = opcionesFindings.get()
    if opcionSoloHigh in ['High CIS','Solo High']:
        while countRow <= sheetCIS.max_row:
            high = sheetCIS.cell(row = countRow, column = referenceSeverityColumn)
            if high.value not in ['severity','High']:
                sheetCIS.delete_rows(countRow)
            else:
                countRow += 1    
      
    countColumn = 1
    while countColumn <= sheetCIS.max_column:
        if sheetCIS.cell(row = 1, column = countColumn).value not in ['complianceStandard', 'complianceControl','complianceControlName','subscriptionName','resourceId','recommendationDisplayName','description','severity','azurePortalRecommendationLink']:
            sheetCIS.delete_cols(countColumn)
        else:
            sheetCIS.cell(row = 1, column = countColumn).value
            countColumn += 1
       
    # ['complianceStandard', 'complianceControl','complianceControlName','subscriptionName','resourceId','recommendationDisplayName','description','severity',azurePortalRecommendationLink']   
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'azurePortalRecommendationLink' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 13 - idxCol)

    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'resourceId' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 12 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'description' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 11 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'severity' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 10 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'complianceControlName' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 9 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'recommendationDisplayName' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 8 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'complianceControl' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 3 - idxCol)   
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'complianceStandard' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 2 - idxCol)   
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetCIS.cell(row = 1, column = idxCol).value == 'subscriptionName' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetCIS.max_row + 1):
                sheetCIS._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 1 - idxCol)     
    
    #Formatear Categoria
    for idxRow in range(2,sheetCIS.max_row + 1):
        if (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '1':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Identity and Access Management'
        elif (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '2':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Security Center'
        elif (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '3':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Storage Accounts'
        elif (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '4':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Database Services'        
        elif (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '5':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Logging and Monitoring'        
        elif (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '6':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Networking'                
        elif (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '7':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Virtual Machines' 
        elif (sheetCIS.cell(row = idxRow, column = 3).value)[0] == '8':
            sheetCIS.cell(row = idxRow, column = 3).value = 'Other Security Considerations'    
        else:
            sheetCIS.cell(row = idxRow, column = 3).value = 'AppService'    
    
    global cisPath
    cisPath = directorypath + '\\CIS.xlsx'
        
    print('End')
    cleanCIS.save(directorypath + '\\CIS.xlsx')
 
dictBM = {}   
def BMDictionary():
    bmDictionary = filedialog.askopenfilename(initialdir = "C:\\Users" ,title = "Seleccionar un archivo", filetypes = (("Archivo Excel", "*.xlsx*"), ("all files", "*.*")))     
    wbDict = openpyxl.load_workbook(os.path.abspath(bmDictionary))
    sheetDict = wbDict.active

    label_file_explorer.configure(text="Diccionario importado: "+ bmDictionary, fg = "#38EB5C" )
   
    global dictBM

    for idxRow in range(2,sheetDict.max_row + 1):
        BMRule = sheetDict.cell(row = idxRow, column = 1).value
        CISCategory = sheetDict.cell(row = idxRow, column = 2).value
        dictBM[BMRule] = CISCategory
    print(bool(dictBM))
            
def browseFilesBM(): 
    
    cellStandard = ''
    
    while(cellStandard != 'Azure Security Benchmark'):
        bm = filedialog.askopenfilename(initialdir = "C:\\Users" ,title = "Seleccionar un archivo", filetypes = (("Archivo CSV", "*.csv*"), ("all files", "*.*")))     
    
        absolutepath = os.path.abspath(bm)
        directorypath = os.path.dirname(absolutepath)
        
        wb = openpyxl.Workbook()
        del wb[wb.sheetnames[0]]
         
        with open(bm, newline='') as f_input:
            ws = wb.create_sheet(title="BM")
            
            for row in csv.reader(f_input, delimiter=','):
                ws.append(row)
                
        wb.save(directorypath + '\\BM.xlsx')
          
        label_file_explorer.configure(text="Archivo abierto: "+ bm, fg = "#38EB5C" )

        checkBM = openpyxl.load_workbook(directorypath + '\\BM.xlsx')
        sheetBM = checkBM.active
        cellStandard = sheetBM.cell(row = 2, column = 2)
        cellStandard = cellStandard.value
        
        if(cellStandard != 'Azure Security Benchmark'):
            label_file_explorer.configure(text='Selecciona el archivo descargado con los findings del Azure Security Benchmark', fg = "#f00")
            
            
    #Formatear
    cleanBM = openpyxl.load_workbook(directorypath + '\\BM.xlsx')
    sheetBM = cleanBM.active
    countRow = 1
    countColumn = 1
    referenceUnhealthyColumn = 1
    referenceSeverityColumn = 1
    
    while countColumn <= sheetBM.max_column:
        if sheetBM.cell(row = 1, column = countColumn).value == 'resourceState':
            referenceUnhealthyColumn = countColumn
        countColumn += 1
    
    countColumn = 1        
    while countColumn <= sheetBM.max_column:
        if sheetBM.cell(row = 1, column = countColumn).value == 'severity':
            referenceSeverityColumn = countColumn
        countColumn += 1
        
    while countRow <= sheetBM.max_row:
        unhealthy = sheetBM.cell(row = countRow, column = referenceUnhealthyColumn)
        if unhealthy.value not in ['resourceState','unhealthy']:
            sheetBM.delete_rows(countRow)
        else:
            countRow += 1
   
    countRow = 1
    opcionSoloHigh = opcionesFindings.get()
    if opcionSoloHigh in ['High BM','Solo High']:
        while countRow <= sheetBM.max_row:
            high = sheetBM.cell(row = countRow, column = referenceSeverityColumn)
            if high.value not in ['severity','High']:
                sheetBM.delete_rows(countRow)
            else:
                countRow += 1    
      
    countColumn = 1
    while countColumn <= sheetBM.max_column:
        if sheetBM.cell(row = 1, column = countColumn).value not in ['complianceStandard', 'complianceControl','complianceControlName','subscriptionName','resourceId','recommendationDisplayName','description','severity','azurePortalRecommendationLink']:
            sheetBM.delete_cols(countColumn)
        else:
            sheetBM.cell(row = 1, column = countColumn).value
            countColumn += 1
                   
    # ['complianceStandard', 'complianceControl','complianceControlName','subscriptionName','resourceId','recommendationDisplayName','description','severity',azurePortalRecommendationLink']   
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'azurePortalRecommendationLink' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 13 - idxCol)

    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'resourceId' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 12 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'description' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 11 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'severity' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 10 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'complianceControlName' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 9 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'recommendationDisplayName' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 8 - idxCol)
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'complianceControl' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 3 - idxCol)   
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'complianceStandard' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 2 - idxCol)   
    
    counter = 0
    for idxCol in range(1, 13):
        if sheetBM.cell(row = 1, column = idxCol).value == 'subscriptionName' and counter == 0:
            counter += 1
            for idxRow in range(1, sheetBM.max_row + 1):
                sheetBM._move_cell(row = idxRow, column=idxCol, row_offset = 0, col_offset = 1 - idxCol)    

    global dictBM
    
    if bool(dictBM) != True:
        dictBM = {'Storage accounts should be migrated to new Azure Resource Manager resources': 'Storage Accounts', 'Virtual machines should be migrated to new Azure Resource Manager resources': 'Virtual Machines', 'Allowlist rules in your adaptive application control policy should be updated': 'Identity and Access Management', 'Adaptive application controls for defining safe applications should be enabled on your machines': 'Virtual Machines', 'Geo-redundant backup should be enabled for Azure Database for MySQL': 'Database Services', 'Azure Backup should be enabled for virtual machines': 'Virtual Machines', 'Microsoft Defender for Storage should be enabled': 'Security Center', 'Microsoft Defender for Azure SQL Database servers should be enabled': 'Security Center', 'Microsoft Defender for SQL servers on machines should be enabled': 'Security Center', 'Microsoft Defender for open-source relational databases should be enabled': 'Security Center', 'Secure transfer to storage accounts should be enabled': 'Storage Accounts', 'FTPS should be required in web apps': 'AppService', 'Web Application should only be accessible over HTTPS': 'AppService', 'TLS should be updated to the latest version for web apps': 'AppService', 'TLS should be updated to the latest version for function apps': 'Networking', 'FTPS should be required in function apps': 'AppService', 'Function App should only be accessible over HTTPS': 'Networking', 'Enforce SSL connection should be enabled for MySQL database servers': 'Database Services', 'Windows web servers should be configured to use secure communication protocols': 'Networking', 'Virtual machines should encrypt temp disks, caches, and data flows between Compute and Storage resources': 'Virtual Machines', 'Transparent Data Encryption on SQL databases should be enabled': 'Database Services', 'Automation account variables should be encrypted': 'Other Security Considerations', 'Microsoft Defender for Key Vault should be enabled': 'Security Center', 'Private endpoint should be configured for Key Vault': 'Networking', 'Key vaults should have purge protection enabled': 'Other Security Considerations', 'Firewall should be enabled on Key Vault': 'Other Security Considerations', 'Key vaults should have soft delete enabled': 'Other Security Considerations', 'Diagnostic logs in Key Vault should be enabled': 'Logging and Monitoring', 'Container hosts should be configured securely': 'Other Security Considerations', 'Microsoft Defender for servers should be enabled': 'Security Center', 'Install endpoint protection solution on virtual machines': 'Virtual Machines', 'Windows Defender Exploit Guard should be enabled on machines': 'Security Center', 'Endpoint protection should be installed on machines': 'Virtual Machines', 'Endpoint protection health issues on virtual machine scale sets should be resolved': 'Virtual Machines', 'SQL servers should have an Azure Active Directory administrator provisioned': 'Database Services', 'Service principals should be used to protect your subscriptions instead of Management Certificates': 'Identity and Access Management', 'Managed identity should be used in web apps': 'Identity and Access Management', 'Managed identity should be used in function apps': 'Identity and Access Management', 'MFA should be enabled on accounts with owner permissions on your subscription': 'Identity and Access Management', 'MFA should be enabled on accounts with write permissions on your subscription': 'Identity and Access Management', 'MFA should be enabled on accounts with read permissions on your subscription': 'Identity and Access Management', 'Authentication to Linux machines should require SSH keys': 'Virtual Machines', 'Email notification for high severity alerts should be enabled': 'Logging and Monitoring', 'Subscriptions should have a contact email address for security issues': 'Logging and Monitoring', 'Email notification to subscription owner for high severity alerts should be enabled': 'Logging and Monitoring', 'Microsoft Defender for App Service should be enabled': 'Security Center', 'Microsoft Defender for DNS should be enabled': 'Security Center', 'Microsoft Defender for Resource Manager should be enabled': 'Security Center', 'Microsoft Defender for SQL should be enabled for unprotected Azure SQL servers': 'Security Center', 'Network Watcher should be enabled': 'Logging and Monitoring', 'Diagnostic logs in App Service should be enabled': 'Logging and Monitoring', 'Diagnostic logs in Event Hub should be enabled': 'Logging and Monitoring', 'Diagnostic logs in Search services should be enabled': 'Logging and Monitoring', 'Auditing on SQL server should be enabled': 'Database Services', 'Diagnostic logs in Azure Stream Analytics should be enabled': 'Logging and Monitoring', 'Diagnostic logs in Service Bus should be enabled': 'Logging and Monitoring', 'Diagnostic logs in Logic Apps should be enabled': 'Logging and Monitoring', 'Diagnostic logs in IoT Hub should be enabled': 'Logging and Monitoring', 'Diagnostic logs in Virtual Machine Scale Sets should be enabled': 'Logging and Monitoring', 'Auto provisioning of the Log Analytics agent should be enabled on subscriptions': 'Logging and Monitoring', 'Log Analytics agent should be installed on virtual machines': 'Logging and Monitoring', 'Log Analytics agent should be installed on virtual machine scale sets': 'Logging and Monitoring', 'Audit retention for SQL servers should be set to at least 90 days': 'Database Services', 'Non-internet-facing virtual machines should be protected with network security groups': 'Networking', 'All network ports should be restricted on network security groups associated to your virtual machine': 'Networking', 'Internet-facing virtual machines should be protected with network security groups': 'Networking', 'Adaptive network hardening recommendations should be applied on internet facing virtual machines': 'Virtual Machines', 'Storage account public access should be disallowed': 'Storage Accounts', 'Storage accounts should restrict network access using virtual network rules': 'Storage Accounts', 'Storage account should use a private link connection': 'Storage Accounts', 'Public network access should be disabled for Cognitive Services accounts': 'Networking', 'Cognitive Services accounts should restrict network access': 'Networking', 'Azure Cosmos DB accounts should have firewall rules': 'Database Services', 'Public network access on Azure SQL Database should be disabled': 'Database Services', 'Private endpoint connections on Azure SQL Database should be enabled': 'Database Services', 'Container registries should use private link': 'Networking', 'Container registries should not allow unrestricted network access': 'Networking', 'Private endpoint should be enabled for MySQL servers': 'Database Services', 'Public network access should be disabled for MySQL servers': 'Database Services', 'Azure Event Grid topics should use private link': 'Networking', 'IP forwarding on your virtual machine should be disabled': 'Networking', 'Management ports of virtual machines should be protected with just-in-time network access control': 'Virtual Machines', 'Management ports should be closed on your virtual machines': 'Virtual Machines', 'Virtual networks should be protected by Azure Firewall': 'Networking', 'Azure DDoS Protection Standard should be enabled': 'Security Center', 'Web Application Firewall (WAF) should be enabled for Azure Front Door Service service': 'Networking', 'Deprecated accounts with owner permissions should be removed from your subscription': 'Identity and Access Management', 'External accounts with owner permissions should be removed from your subscription': 'Identity and Access Management', 'There should be more than one owner assigned to subscriptions': 'Identity and Access Management', 'A maximum of 3 owners should be designated for subscriptions': 'Identity and Access Management', 'External accounts with write permissions should be removed from your subscription': 'Identity and Access Management', 'Deprecated accounts should be removed from your subscription': 'Identity and Access Management', 'External accounts with read permissions should be removed from your subscription': 'Identity and Access Management', 'Remote debugging should be turned off for Web Applications': 'AppService', 'Web apps should request an SSL certificate for all incoming requests': 'AppService', 'CORS should not allow every resource to access Web Applications': 'AppService', 'Remote debugging should be turned off for Function App': 'Other Security Considerations', 'CORS should not allow every resource to access Function Apps': 'Other Security Considerations', 'Function apps should have Client Certificates (Incoming client certificates) enabled': 'Identity and Access Management', 'Vulnerabilities in security configuration on your Windows machines should be remediated (powered by Guest Configuration)': 'Virtual Machines', 'Vulnerabilities in security configuration on your Linux machines should be remediated (powered by Guest Configuration)': 'Virtual Machines', 'Secure Boot should be enabled on supported Windows virtual machines': 'Virtual Machines', 'Guest Configuration extension should be installed on machines': 'Virtual Machines', 'vTPM should be enabled on supported virtual machines': 'Virtual Machines', 'Machines should have vulnerability findings resolved': 'Virtual Machines', 'SQL servers should have vulnerability assessment configured': 'Database Services', 'Machines should be configured securely': 'Virtual Machines', 'SQL databases should have vulnerability findings resolved': 'Database Services', 'Java should be updated to the latest version for web apps': 'AppService', 'PHP should be updated to the latest version for web apps': 'AppService', 'Python should be updated to the latest version for web apps': 'AppService', 'System updates should be installed on your machines': 'Virtual Machines', 'Virtual machine scale sets should be configured securely': 'Virtual Machines', 'System updates on virtual machine scale sets should be installed': 'Virtual Machines', 'Container registry images should have vulnerability findings resolved': 'Other Security Considerations', 'Python should be updated to the latest version for function apps': 'AppService', 'Azure Machine Learning workspaces should use private link': 'Networking', 'Subnets should be associated with a network security group': 'Networking', "Virtual machines' Guest Configuration extension should be deployed with system-assigned managed identity": 'Virtual Machines', 'Java should be updated to the latest version for function apps': 'AppService', 'Endpoint protection health issues on machines should be resolved': 'Virtual Machines'}
    
    #Formatear Categoria
    for idxRow in range(2,sheetBM.max_row + 1):
        if sheetBM.cell(row = idxRow, column = 8).value in dictBM.keys():
            sheetBM.cell(row = idxRow, column = 3).value = dictBM[sheetBM.cell(row = idxRow, column = 8).value]
        else:
            sheetBM.cell(row = idxRow, column = 3).value = "Sin definir"
    
    global bmPath
    bmPath = directorypath + '\\BM.xlsx'
            
    print('End')
    cleanBM.save(directorypath + '\\BM.xlsx')

def posible_Subscriptions():
    global subscription
    subscription = posibleSubscriptions.get()
    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription] 
    referenceSubColumn = 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Account Name':
            referenceSubColumn = countColumn
        countColumn += 1

    subscriptionsInPart = []
    for idxRow in range(2, sheetCSPM.max_row + 1):
        if sheetCSPM.cell(row = idxRow, column = referenceSubColumn).value not in subscriptionsInPart and sheetCSPM.cell(row = idxRow, column = referenceSubColumn).value != None:
            subscriptionsInPart.append(sheetCSPM.cell(row = idxRow, column = referenceSubColumn).value)
    
    if len(subscriptionsInPart) > 0:
        if len(subscriptionsInPart) > 1:
            subscriptionsInPart.insert(0, 'Todas las suscripciones')
            posibleSubscriptionsParticular.configure(values = subscriptionsInPart)
            posibleSubscriptionsParticular.current(0)
        else:
            posibleSubscriptionsParticular.configure(values = subscriptionsInPart)
            posibleSubscriptionsParticular.current(0)
    
    print(posibleSubscriptions.get())

def isSelected():
    if posibleSubscriptions.get() == None:
        return False
    else:
        return True
    
#####################################################EVALUACION##########################################################

def compareBM2CIS(): #PARA OPCION TODAS
    wbCIS = openpyxl.load_workbook(cisPath)
    wbBM = openpyxl.load_workbook(bmPath)
    sheetCIS = wbCIS.active
    sheetBM = wbBM.active
    delAlwaysBM = ['MFA should be enabled on accounts with owner permissions on your subscription',
    'MFA should be enabled on accounts with write permissions on your subscription',
    'MFA should be enabled on accounts with read permissions on your subscription',
    'External accounts with write permissions should be removed from your subscription',
    'External accounts with owner permissions should be removed from your subscription',
    'External accounts with read permissions should be removed from your subscription']
    
    for idxRowCIS in range (2, sheetCIS.max_row + 1):
        idxRowBM = 2
        while (idxRowBM < sheetBM.max_row + 1):
            if sheetBM.cell(row = idxRowBM, column = 8).value in delAlwaysBM or (sheetCIS.cell(row = idxRowCIS, column = 8).value == sheetBM.cell(row = idxRowBM, column = 8).value and sheetCIS.cell(row = idxRowCIS, column = 12).value == sheetBM.cell(row = idxRowBM, column = 12).value):    
                sheetBM.delete_rows(idxRowBM)
                wbBM.save(bmPath)
            else:
                idxRowBM += 1
    
    
           
def compareIfPendings():
    wbCSPM = openpyxl.load_workbook(cspmPath)
    if os.path.exists(cisPath):
        wbCIS = openpyxl.load_workbook(cisPath)
        sheetCIS = wbCIS.active
    if os.path.exists(bmPath):
        wbBM = openpyxl.load_workbook(bmPath)
        sheetBM = wbBM.active

    global subscription
    print(subscription)
    
    while not isSelected():
        print("Select a subscription")
        sleep(10)
        
    sheetCSPM = wbCSPM[subscription]
    
    addAlwaysCIS = ['MFA should be enabled on accounts with write permissions on subscriptions',
    'MFA should be enabled on accounts with read permissions on subscriptions',
   ' MFA should be enabled on accounts with owner permissions on subscriptions',
    'External accounts with owner permissions should be removed from subscriptions',
    'External accounts with read permissions should be removed from subscriptions',
    'External accounts with write permissions should be removed from subscriptions']
    addAlwaysBM = ['MFA should be enabled on accounts with owner permissions on your subscription',
    'MFA should be enabled on accounts with write permissions on your subscription',
    'MFA should be enabled on accounts with read permissions on your subscription',
    'External accounts with write permissions should be removed from your subscription',
    'External accounts with owner permissions should be removed from your subscription',
    'External accounts with read permissions should be removed from your subscription']

    #New to Pending (column 7)
    for idxRow in range (2, sheetCSPM.max_row +1):
        if sheetCSPM.cell(row = idxRow, column = 7).value == 'New':
            sheetCSPM.cell(row = idxRow, column = 7).value = 'Pending'
       
    #Compare
    
    ##Only CIS
    if os.path.exists(cisPath) and not os.path.exists(bmPath):
        maxCSPMInitialRows = 0
        for idxRow in range(1, sheetCSPM.max_row + 1):
            if sheetCSPM.cell(row = idxRow, column = 1).value != None:
                maxCSPMInitialRows += 1
        maxInitialRowsCSPM = sheetCSPM.max_row #No se actualiza, excel inicial sin novedades
        maxRowsCSPM = maxInitialRowsCSPM + 1 #Se actualiza para las filas nuevas
        for idxRowCIS in range (2, sheetCIS.max_row + 1):
            if sheetCIS.cell(row = idxRowCIS, column = 8).value not in addAlwaysCIS:
                isPending = False 
                for idxRowCSPM in range (2, maxInitialRowsCSPM + 1):
                    if sheetCIS.cell(row = idxRowCIS, column = 1).value == sheetCSPM.cell(row = idxRowCSPM, column = 1).value and sheetCSPM.cell(row = idxRowCSPM, column = 7).value not in ["Done","New"]: #FINDING MISMA SUSCRIPCION Y NO RESUELTOS (EVITAR FALSOS POSITIVOS)
                        if sheetCIS.cell(row = idxRowCIS, column = 8).value == sheetCSPM.cell(row = idxRowCSPM, column = 8).value and sheetCIS.cell(row = idxRowCIS, column = 9).value == sheetCSPM.cell(row = idxRowCSPM, column = 9).value and sheetCIS.cell(row = idxRowCIS, column = 12).value == sheetCSPM.cell(row = idxRowCSPM, column = 12).value:
                            isPending = True
                if isPending == False:
                    for idxCol in range(1,14):
                        sheetCSPM.cell(row = maxRowsCSPM, column = idxCol).value = sheetCIS.cell(row = idxRowCIS, column = idxCol).value
                    sheetCSPM.cell(row = maxRowsCSPM, column = 7).value = "New"
                    maxRowsCSPM += 1
            else: #ADD MFAS AND EXTERNAL ACCOUNTS
                for idxCol in range(1,14):
                    sheetCSPM.cell(row = maxRowsCSPM, column = idxCol).value = sheetCIS.cell(row = idxRowCIS, column = idxCol).value
                sheetCSPM.cell(row = maxRowsCSPM, column = 7).value = "New"
                maxRowsCSPM += 1                                 
    
    ##Only BM
    elif os.path.exists(bmPath) and not os.path.exists(cisPath):        
        maxCSPMInitialRows = 0
        for idxRow in range(1, sheetCSPM.max_row + 1):
            if sheetCSPM.cell(row = idxRow, column = 1).value != None:
                maxCSPMInitialRows += 1 
        maxInitialRowsCSPM = sheetCSPM.max_row #No se actualiza, excel inicial sin novedades
        maxRowsCSPM = maxInitialRowsCSPM + 1 #Se actualiza para las filas nuevas
        for idxRowBM in range (2, sheetBM.max_row + 1):
            if sheetBM.cell(row = idxRowBM, column = 8).value not in addAlwaysBM:
                isPending = False
                for idxRowCSPM in range (2, maxInitialRowsCSPM + 1):
                    if sheetBM.cell(row = idxRowBM, column = 1).value == sheetCSPM.cell(row = idxRowCSPM, column = 1).value and sheetCSPM.cell(row = idxRowCSPM, column = 7).value not in ["Done","New"]: #FINDING MISMA SUSCRIPCION Y NO RESUELTOS (EVITAR FALSOS POSITIVOS)
                        if sheetBM.cell(row = idxRowBM, column = 8).value == sheetCSPM.cell(row = idxRowCSPM, column = 8).value and sheetBM.cell(row = idxRowBM, column = 9).value == sheetCSPM.cell(row = idxRowCSPM, column = 9).value and sheetBM.cell(row = idxRowBM, column = 12).value == sheetCSPM.cell(row = idxRowCSPM, column = 12).value:
                            isPending = True
                if isPending == False:
                    for idxCol in range(1,14):
                        sheetCSPM.cell(row = maxRowsCSPM, column = idxCol).value = sheetBM.cell(row = idxRowBM, column = idxCol).value
                    sheetCSPM.cell(row = maxRowsCSPM, column = 7).value = "New"
                    maxRowsCSPM += 1
            else: #ADD MFAS AND EXTERNAL ACCOUNTS
                for idxCol in range(1,14):
                    sheetCSPM.cell(row = maxRowsCSPM, column = idxCol).value = sheetBM.cell(row = idxRowBM, column = idxCol).value
                sheetCSPM.cell(row = maxRowsCSPM, column = 7).value = "New"
                maxRowsCSPM += 1                                  

    ## Both submitted
    elif os.path.exists(cisPath) and os.path.exists(bmPath):
        compareBM2CIS() #ELIMINAR FILAS BM IGUALES QUE CIS

        wbBM = openpyxl.load_workbook(bmPath)
        sheetBM = wbBM.active        
        #COMPARE TO CIS
        maxCSPMInitialRows = 0
        for idxRow in range(1, sheetCSPM.max_row + 1):
            if sheetCSPM.cell(row = idxRow, column = 1).value != None:
                maxCSPMInitialRows += 1
        maxInitialRowsCSPM = sheetCSPM.max_row #No se actualiza, excel inicial sin novedades
        maxRowsCSPM = maxInitialRowsCSPM + 1 #Se actualiza para las filas nuevas
        for idxRowCIS in range (2, sheetCIS.max_row + 1):
            if sheetCIS.cell(row = idxRowCIS, column = 8).value not in addAlwaysCIS:
                isPending = False 
                for idxRowCSPM in range (2, maxInitialRowsCSPM + 1):
                    if sheetCIS.cell(row = idxRowCIS, column = 1).value == sheetCSPM.cell(row = idxRowCSPM, column = 1).value and sheetCSPM.cell(row = idxRowCSPM, column = 7).value not in ["Done","New"]: #FINDING MISMA SUSCRIPCION Y NO RESUELTOS (EVITAR FALSOS POSITIVOS)
                        if sheetCIS.cell(row = idxRowCIS, column = 8).value == sheetCSPM.cell(row = idxRowCSPM, column = 8).value and sheetCIS.cell(row = idxRowCIS, column = 9).value == sheetCSPM.cell(row = idxRowCSPM, column = 9).value and sheetCIS.cell(row = idxRowCIS, column = 12).value == sheetCSPM.cell(row = idxRowCSPM, column = 12).value:
                            isPending = True
                if isPending == False:
                    for idxCol in range(1,14):
                        sheetCSPM.cell(row = maxRowsCSPM, column = idxCol).value = sheetCIS.cell(row = idxRowCIS, column = idxCol).value
                    sheetCSPM.cell(row = maxRowsCSPM, column = 7).value = "New"
                    maxRowsCSPM += 1
            else: #ADD MFAS AND EXTERNAL ACCOUNTS
                for idxCol in range(1,14):
                    sheetCSPM.cell(row = maxRowsCSPM, column = idxCol).value = sheetCIS.cell(row = idxRowCIS, column = idxCol).value
                sheetCSPM.cell(row = maxRowsCSPM, column = 7).value = "New"
                maxRowsCSPM += 1  
                
        #COMPARE TO BM
        for idxRow in range(1, sheetCSPM.max_row + 1):
            if sheetCSPM.cell(row = idxRow, column = 1).value != None:
                maxCSPMInitialRows += 1 
        maxInitialRowsCSPM = sheetCSPM.max_row #No se actualiza, excel inicial sin novedades
        maxRowsCSPM = maxInitialRowsCSPM + 1 #Se actualiza para las filas nuevas
        for idxRowBM in range (2, sheetBM.max_row + 1):
            isPending = False
            for idxRowCSPM in range (2, maxInitialRowsCSPM + 1):
                if sheetBM.cell(row = idxRowBM, column = 1).value == sheetCSPM.cell(row = idxRowCSPM, column = 1).value and sheetCSPM.cell(row = idxRowCSPM, column = 7).value not in ["Done","New"]: #FINDING MISMA SUSCRIPCION Y NO RESUELTOS (EVITAR FALSOS POSITIVOS)
                    if sheetBM.cell(row = idxRowBM, column = 8).value == sheetCSPM.cell(row = idxRowCSPM, column = 8).value and sheetBM.cell(row = idxRowBM, column = 9).value == sheetCSPM.cell(row = idxRowCSPM, column = 9).value and sheetBM.cell(row = idxRowBM, column = 12).value == sheetCSPM.cell(row = idxRowCSPM, column = 12).value:
                        isPending = True
            if isPending == False:
                for idxCol in range(1,14):
                    sheetCSPM.cell(row = maxRowsCSPM, column = idxCol).value = sheetBM.cell(row = idxRowBM, column = idxCol).value
                sheetCSPM.cell(row = maxRowsCSPM, column = 7).value = "New"
                maxRowsCSPM += 1
                  
    else:
        print("Ningun archivo del portal introducido")
        
    print('End')
    wbCSPM.save(cspmPath)
    label_file_explorer.configure(text="Evaluacion finalizada ", fg = "#38EB5C" )

#####################################################ESTABLECER COMO RESUELTO##########################################################

suscripcionElegida = ''
posibleSuscripcion = ''
listadofindings = []
listadorecursos = []
 
def suscripcionSeleccionada():
    global posiblesSuscripciones
    global suscripcionElegida
    global posibleFindings
    global listadofindings
    global cspmPath

    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription]

    suscripcionElegida = posiblesSuscripciones.get()
    print(posiblesSuscripciones.get())
    listadofindings = []
    for idxRow in range(1, sheetCSPM.max_row + 1):
        if sheetCSPM.cell(row = idxRow, column = 1).value == suscripcionElegida:
            if sheetCSPM.cell(row = idxRow, column = 7).value != 'Done' and sheetCSPM.cell(row = idxRow, column = 8).value not in listadofindings and sheetCSPM.cell(row = idxRow, column = 8).value != 'Rule Title':
                listadofindings.append(sheetCSPM.cell(row = idxRow, column = 8).value) 
    posibleFindings.configure(values = listadofindings)

posibleFinding = ''
def findingSeleccionado():
    global posiblesSuscripciones
    global suscripcionElegida
    global posibleFinding
    global posibleFindings
    global posiblesRecursos
    global listadorecursos

    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription]

    posibleFinding = posibleFindings.get()
    print(posibleFindings.get())
    listadorecursos = []
    for idxRow in range(1, sheetCSPM.max_row + 1):
        if sheetCSPM.cell(row = idxRow, column = 1).value == suscripcionElegida and sheetCSPM.cell(row = idxRow, column = 8).value == posibleFinding:
            if sheetCSPM.cell(row = idxRow, column = 7).value != 'Done' and sheetCSPM.cell(row = idxRow, column = 12).value not in listadofindings and sheetCSPM.cell(row = idxRow, column = 12).value != 'Resource':
                listadorecursos.append(sheetCSPM.cell(row = idxRow, column = 12).value) 
    posiblesRecursos.configure(values = listadorecursos)   
    
def recursoResuelto():
    global suscripcionElegida
    global posibleFinding
    global posibleFindings
    global listadofindings
    global posiblesRecursos
    global listadorecursos

    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription]

    resuelto = posiblesRecursos.get()
    print(posiblesRecursos.get())
    for idxRow in range(1, sheetCSPM.max_row + 1):
        if sheetCSPM.cell(row = idxRow, column = 1).value == suscripcionElegida and sheetCSPM.cell(row = idxRow, column = 8).value == posibleFinding and sheetCSPM.cell(row = idxRow, column = 12).value == resuelto:
            sheetCSPM.cell(row = idxRow, column = 7).value = 'Done'
            wbCSPM.save(cspmPath)
    listadorecursos.remove(resuelto)
    posiblesRecursos.configure(values = listadorecursos)   
    if(len(listadorecursos) == 0):
        listadofindings.remove(posibleFinding)
        posibleFindings.configure(values = listadofindings)     

def recursoResueltoPorPosicion():
    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription]
    
    global posiblesRecursosPorPosicion
    
    idx = posiblesRecursosPorPosicion.get()
    print('resolver el del index: ' + idx)
    idx = int(idx)
    print('Actualmente: ' + sheetCSPM.cell(row = idx, column = 7).value)
    if sheetCSPM.cell(row = idx, column = 7).value != None and sheetCSPM.cell(row = idx, column = 7).value != 'Situation':  
        sheetCSPM.cell(row = idx, column = 7).value = 'Done'  
    print('Ahora: ' + sheetCSPM.cell(row = idx, column = 7).value)
    wbCSPM.save(cspmPath)
    
def setToDone():
    wbCSPM = openpyxl.load_workbook(cspmPath)
    
    while not isSelected():
        print("Select a subscription")
        sleep(10)
    
    sheetCSPM = wbCSPM[subscription]
    listadoDeSuscripciones = []
    
    for idxRow in range(1, sheetCSPM.max_row + 1):
        if sheetCSPM.cell(row = idxRow, column = 1).value != None and sheetCSPM.cell(row = idxRow, column = 1).value != 'Account Name':
            if sheetCSPM.cell(row = idxRow, column = 1).value not in listadoDeSuscripciones:
                listadoDeSuscripciones.append(sheetCSPM.cell(row = idxRow, column = 1).value)
    print(listadoDeSuscripciones)
    
    listadoDeFindingsNoResueltos = []
    for idxRow in range(1, sheetCSPM.max_row + 1):
        if sheetCSPM.cell(row = idxRow, column = 1).value != None and sheetCSPM.cell(row = idxRow, column = 7).value != 'Situation' and sheetCSPM.cell(row = idxRow, column = 7).value != 'Done':
            listadoDeFindingsNoResueltos.append(idxRow)
   
    
    done = Toplevel(window)

    done.title("Establecer como resueltos")
    done.geometry("1500x500")

    global posiblesSuscripciones
    global suscripcionElegida
    global posibleFindings
    global listadofindings
    global posiblesRecursos
    global listadorecursos
    global posiblesRecursosPorPosicion
    
    opciona = Label(done,  text = "Opcion 1")
    opciona.grid(row = 0, column = 0, padx=10, pady=10, sticky="nsew") 

    suscripcionAElegir = Label(done,  text = "Elegir suscripcion")
    suscripcionAElegir.grid(row = 2, column = 0, padx=10, pady=10, sticky="nsew") 
    posiblesSuscripciones = ttk.Combobox(done,state="readonly",values=listadoDeSuscripciones)
    posiblesSuscripciones.grid(row = 3, column = 0, padx=10, pady=10, sticky="nsew") 
    
    guardarSus = Button(done, text = "Seleccionar suscripcion", command = suscripcionSeleccionada)
    guardarSus.grid(row = 3, column = 1, padx=10, pady=10, sticky="nsew")  
    
    findingAElegir = Label(done,  text = "Elegir finding")
    findingAElegir.grid(row = 4, column = 0, padx=10, pady=10, sticky="nsew") 
    posibleFindings = ttk.Combobox(done,width = 150,state="readonly",values=listadofindings)
    posibleFindings.grid(row = 5, column = 0, padx=10, pady=10, sticky="nsew") 
    
    guardarFind = Button(done, text = "Seleccionar finding", command = findingSeleccionado)
    guardarFind.grid(row = 5, column = 1, padx=10, pady=10, sticky="nsew") 
    
    recursoAElegir = Label(done,  text = "Elegir recurso")
    recursoAElegir.grid(row = 6, column = 0, padx=10, pady=10, sticky="nsew") 
    posiblesRecursos = ttk.Combobox(done,width = 200,state="readonly",values=listadorecursos)
    posiblesRecursos.grid(row = 7, column = 0, padx=10, pady=10, sticky="nsew") 
    
    recursoADone = Button(done, text = "Recurso resuelto", command = recursoResuelto)
    recursoADone.grid(row = 7, column = 1, padx=10, pady=10, sticky="nsew") 

    opcionb = Label(done,  text = "-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
    opcionb.grid(row = 8, columnspan = 2, padx=10, pady=10, sticky="nsew") 
    
    opcionb = Label(done,  text = "Opcion 2")
    opcionb.grid(row = 9, column = 0, padx=10, pady=10, sticky="nsew") 
    
    recursoAElegirPorPosicion = Label(done,  text = "Posicion en Excel")
    recursoAElegirPorPosicion.grid(row = 11, column = 0, padx=10, pady=10, sticky="nsew") 
    posiblesRecursosPorPosicion = Entry(done)
    posiblesRecursosPorPosicion.grid(row = 12, column = 0, padx=10, pady=10, sticky="nsew")     
    
    recursoADoneporPosicion = Button(done, text = "Recurso resuelto", command = recursoResueltoPorPosicion)
    recursoADoneporPosicion.grid(row = 12, column = 1, padx=10, pady=10, sticky="nsew") 
    


#####################################################RESULTADOS##########################################################

listaConteoRL = [] 
def conteoRiskLevel():
    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription] 
    referenceSeverityColumn = 1
    referenceSituationColumn = 1
    referenceSubColumn = 1
    high = 0
    medium = 0
    low = 0
    countColumn = 1
    subIndividual = posibleSubscriptionsParticular.get()
    print('Sub en particular: ' + subIndividual)
    global listaConteoRL
    listaConteoRL = []
    
    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Risk Level':
            referenceSeverityColumn = countColumn
        countColumn += 1
    
    countColumn = 1        
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Situation':
            referenceSituationColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Account Name':
            referenceSubColumn = countColumn
        countColumn += 1
    print('Maximo de filas: ' + str(sheetCSPM.max_row))        
    if subIndividual == 'Todas las suscripciones':       
        for idxRow in range(2, sheetCSPM.max_row + 1):
            if sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value in ['Pending','New']:
                riskLevel = sheetCSPM.cell(row = idxRow, column = referenceSeverityColumn).value
               
                if riskLevel == 'High':
                    high += 1
                elif riskLevel == 'Medium':
                    medium += 1
                else:
                    low += 1
    
    else:
        for idxRow in range(2, sheetCSPM.max_row + 1):
            if sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value in ['Pending','New'] and sheetCSPM.cell(row = idxRow, column = referenceSubColumn).value == subIndividual:
                riskLevel = sheetCSPM.cell(row = idxRow, column = referenceSeverityColumn).value
               
                if riskLevel == 'High':
                    high += 1
                elif riskLevel == 'Medium':
                    medium += 1
                else:
                    low += 1        
    
    listaConteoRL = [['Severity', '# Recursos'],['High', high],['Medium',medium],['Low',low],]
    print('High: ' + str(high))
    print('Medium: ' + str(medium))
    print('Low: ' + str(low))
    print(sheetCSPM.max_row + 1)

listaConteoCat = []
def conteoCatCIS():
    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription] 
    referenceCategoryColumn = 1
    referenceSituationColumn = 1
    referenceSubColumn = 1
    IAM = 0
    SC = 0
    SA = 0
    DS = 0
    LM = 0
    N = 0
    VM = 0
    OSC = 0
    AS = 0
    countColumn = 1
    subIndividual = posibleSubscriptionsParticular.get()
    global listaConteoCat
    listaConteoCat = []
    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Category':
            referenceCategoryColumn = countColumn
        countColumn += 1
        
    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Situation':
            referenceSituationColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Account Name':
            referenceSubColumn = countColumn
        countColumn += 1
    
    if subIndividual == 'Todas las suscripciones':           
        for idxRow in range(2, sheetCSPM.max_row + 1):   
            if sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value in ['Pending','New']:
                category = sheetCSPM.cell(row = idxRow, column = referenceCategoryColumn).value
           
                if category == 'Identity and Access Management':
                    IAM += 1
                elif category == 'Security Center':
                    SC += 1
                elif category == 'Storage Accounts':
                    SA += 1
                elif category == 'Database Services':
                    DS += 1
                elif category == 'Logging and Monitoring':
                    LM += 1
                elif category ==  'Networking':
                    N += 1
                elif category == 'Virtual Machines':
                    VM += 1 
                elif category ==  'Other Security Considerations':
                    OSC +=1
                else:
                    AS += 1
    else: 
        for idxRow in range(2, sheetCSPM.max_row + 1):   
            if sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value in ['Pending','New'] and sheetCSPM.cell(row = idxRow, column = referenceSubColumn).value == subIndividual:
                category = sheetCSPM.cell(row = idxRow, column = referenceCategoryColumn).value
           
                if category == 'Identity and Access Management':
                    IAM += 1
                elif category == 'Security Center':
                    SC += 1
                elif category == 'Storage Accounts':
                    SA += 1
                elif category == 'Database Services':
                    DS += 1
                elif category == 'Logging and Monitoring':
                    LM += 1
                elif category ==  'Networking':
                    N += 1
                elif category == 'Virtual Machines':
                    VM += 1 
                elif category ==  'Other Security Considerations':
                    OSC +=1
                else:
                    AS += 1        
    
    listaConteoCat = [['Category', '# Recursos'],
                    ['Identity and Access Management', IAM],
                    ['Security Center',SC],
                    ['Storage Accounts',SA],
                    ['Database Services',DS],
                    ['Logging and Monitoring',LM],
                    ['Networking',N],
                    ['Virtual Machines',VM],
                    ['Other Security Considerations',OSC], 
                    ['AppService',AS],                    
    ]
    print(len(listaConteoCat))
    for x in range(0,len(listaConteoCat)):
        print(listaConteoCat[x][0] + ': ' + str(listaConteoCat[x][1]))

listaConteoRLpCat = []
def conteoRLpCatCIS():
    wbCSPM = openpyxl.load_workbook(cspmPath)
    sheetCSPM = wbCSPM[subscription] 
    referenceCategoryColumn = 1
    referenceSituationColumn = 1
    referenceSubColumn = 1
    referenceSeverityColumn = 1

    IAM = ['Identity and Access Management', 0,0,0]
    SC = ['Security Center',0,0,0]
    SA = ['Storage Accounts',0,0,0]
    DS = ['Database Services',0,0,0]
    LM = ['Logging and Monitoring',0,0,0]
    N = ['Networking',0,0,0]
    VM = ['Virtual Machines',0,0,0]
    OSC = ['Other Security Considerations',0,0,0]
    AS = ['AppService',0,0,0]
    countColumn = 1
    subIndividual = posibleSubscriptionsParticular.get()
    global listaConteoRLpCat
    listaConteoRLpCat = [['Category', 'High', 'Medium','Low'],]
    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Category':
            referenceCategoryColumn = countColumn
        countColumn += 1
        
    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Situation':
            referenceSituationColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Account Name':
            referenceSubColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Risk Level':
            referenceSeverityColumn = countColumn
        countColumn += 1
    
    if subIndividual == 'Todas las suscripciones':           
        for idxRow in range(2, sheetCSPM.max_row + 1):   
            if sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value in ['Pending','New']:
                category = sheetCSPM.cell(row = idxRow, column = referenceCategoryColumn).value
                risk = sheetCSPM.cell(row = idxRow, column = referenceSeverityColumn).value
                
                if category == 'Identity and Access Management':
                    if risk == 'High':
                        IAM[1] += 1
                    elif risk == 'Medium':
                        IAM[2] += 1
                    else:
                        IAM[3] += 1
                        
                elif category == 'Security Center':
                    if risk == 'High':
                        SC[1] += 1
                    elif risk == 'Medium':
                        SC[2] += 1
                    else:
                        SC[3] += 1
                elif category == 'Storage Accounts':
                    if risk == 'High':
                        SA[1] += 1
                    elif risk == 'Medium':
                        SA[2] += 1
                    else:
                        SA[3] += 1
                elif category == 'Database Services':
                    if risk == 'High':
                        DS[1] += 1
                    elif risk == 'Medium':
                        DS[2] += 1
                    else:
                        DS[3] += 1
                        
                elif category == 'Logging and Monitoring':
                    if risk == 'High':
                        LM[1] += 1
                    elif risk == 'Medium':
                        LM[2] += 1
                    else:
                        LM[3] += 1
                        
                elif category ==  'Networking':
                    if risk == 'High':
                        N[1] += 1
                    elif risk == 'Medium':
                        N[2] += 1
                    else:
                        N[3] += 1
                     
                elif category == 'Virtual Machines':
                    if risk == 'High':
                        VM[1] += 1
                    elif risk == 'Medium':
                        VM[2] += 1
                    else:
                        VM[3] += 1
                      
                elif category ==  'Other Security Considerations':
                    if risk == 'High':
                        OSC[1] += 1
                    elif risk == 'Medium':
                        OSC[2] += 1
                    else:
                        OSC[3] += 1
                     
                else:
                    if risk == 'High':
                        AS[1] += 1
                    elif risk == 'Medium':
                        AS[2] += 1
                    else:
                        AS[3] += 1
    else: 
        for idxRow in range(2, sheetCSPM.max_row + 1):   
            if sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value in ['Pending','New'] and sheetCSPM.cell(row = idxRow, column = referenceSubColumn).value == subIndividual:
                category = sheetCSPM.cell(row = idxRow, column = referenceCategoryColumn).value
                risk = sheetCSPM.cell(row = idxRow, column = referenceSeverityColumn).value
          
                if category == 'Identity and Access Management':
                    if risk == 'High':
                        IAM[1] += 1
                    elif risk == 'Medium':
                        IAM[2] += 1
                    else:
                        IAM[3] += 1
                        
                elif category == 'Security Center':
                    if risk == 'High':
                        SC[1] += 1
                    elif risk == 'Medium':
                        SC[2] += 1
                    else:
                        SC[3] += 1
                        
                elif category == 'Storage Accounts':
                    if risk == 'High':
                        SA[1] += 1
                    elif risk == 'Medium':
                        SA[2] += 1
                    else:
                        SA[3] += 1
                        
                elif category == 'Database Services':
                    if risk == 'High':
                        DS[1] += 1
                    elif risk == 'Medium':
                        DS[2] += 1
                    else:
                        DS[3] += 1
                        
                elif category == 'Logging and Monitoring':
                    if risk == 'High':
                        LM[1] += 1
                    elif risk == 'Medium':
                        LM[2] += 1
                    else:
                        LM[3] += 1
                     
                elif category ==  'Networking':
                    if risk == 'High':
                        N[1] += 1
                    elif risk == 'Medium':
                        N[2] += 1
                    else:
                        N[3] += 1
                     
                elif category == 'Virtual Machines':
                    if risk == 'High':
                        VM[1] += 1
                    elif risk == 'Medium':
                        VM[2] += 1
                    else:
                        VM[3] += 1
                      
                elif category ==  'Other Security Considerations':
                    if risk == 'High':
                        OSC[1] += 1
                    elif risk == 'Medium':
                        OSC[2] += 1
                    else:
                        OSC[3] += 1
                     
                else:
                    if risk == 'High':
                        AS[1] += 1
                    elif risk == 'Medium':
                        AS[2] += 1
                    else:
                        AS[3] += 1        
    
    listaConteoRLpCat.append(IAM)
    listaConteoRLpCat.append(SC)
    listaConteoRLpCat.append(SA)
    listaConteoRLpCat.append(DS)
    listaConteoRLpCat.append(LM)
    listaConteoRLpCat.append(N)
    listaConteoRLpCat.append(VM)
    listaConteoRLpCat.append(OSC)
    listaConteoRLpCat.append(AS)


def dibujarGRAFS():

    wb = openpyxl.load_workbook(cspmPath)     
    subIndividual = posibleSubscriptionsParticular.get()
    
    if subIndividual == 'Todas las suscripciones':
        titleGraf = 'Grafs ' + str(subscription) 

        if titleGraf not in wb.sheetnames:
            positionS = wb.worksheets.index(wb[subscription]) + 1
            wb.create_sheet(index = positionS, title = titleGraf)
            sheet = wb[titleGraf]
        else: 
            del wb[titleGraf]
            
            positionS = wb.worksheets.index(wb[subscription]) + 1
            wb.create_sheet(index = positionS, title = titleGraf)
            sheet = wb[titleGraf]
    else: 
        titleGraf = 'Grafs ' + str(subIndividual)
        if titleGraf not in wb.sheetnames:
            positionS = wb.worksheets.index(wb[subscription]) + 1
            wb.create_sheet(index = positionS, title = titleGraf)
            sheet = wb[titleGraf]        
        else: 
            del wb[titleGraf]
            
            positionS = wb.worksheets.index(wb[subscription]) + 1
            wb.create_sheet(index = positionS, title = titleGraf)
            sheet = wb[titleGraf]       

    
    conteoRiskLevel()
    global listaConteoRL
    for row in listaConteoRL:
        sheet.append(row)
    
    chartSeverity = PieChart()
    
    labelsSev = Reference(sheet, min_col = 1,min_row = 2, max_row = 4)
    dataSev = Reference(sheet, min_col = 2,min_row = 1, max_row = 4)

    chartSeverity.add_data(dataSev, titles_from_data = True)
    chartSeverity.set_categories(labelsSev)
    chartSeverity.title = "Criticidad"
    
    
    seriesSev = chartSeverity.series[0]
    for i in range(len(listaConteoRL)-1):
        if listaConteoRL[i+1][0] == 'High':
            pt = openpyxl.chart.marker.DataPoint(idx=i)
            pt.graphicalProperties.line.solidFill = 'FF0000'
            pt.graphicalProperties.solidFill = 'FF0000'
            seriesSev.dPt.append(pt)
        if listaConteoRL[i+1][0] == 'Medium':
            pt = openpyxl.chart.marker.DataPoint(idx=i)
            pt.graphicalProperties.line.solidFill = 'FF8000'
            pt.graphicalProperties.solidFill = 'FF8000'
            seriesSev.dPt.append(pt)
        if listaConteoRL[i+1][0] == 'Low':
            pt = openpyxl.chart.marker.DataPoint(idx=i)
            pt.graphicalProperties.line.solidFill = 'FFFF00'
            pt.graphicalProperties.solidFill = 'FFFF00'
            seriesSev.dPt.append(pt)

    chartSeverity.dataLabels = DataLabelList() 
    chartSeverity.dataLabels.showPercent = True
    sheet.add_chart(chartSeverity, "E2")
    
   ################################################### 
    sheet.append(['',''])    
    ################################################### 
   
    conteoCatCIS()
    global listaConteoCat
    for idx in listaConteoCat:
        sheet.append(idx)
    
    
    chartCat = PieChart()
    
    labelsCat = Reference(sheet, min_col = 1,min_row = 7, max_row = 15)
    dataCat = Reference(sheet, min_col = 2,min_row = 7, max_row = 15)   
    
    chartCat.add_data(dataCat, titles_from_data = True)
    chartCat.set_categories(labelsCat)
    chartCat.title = "Categorias"  
    
    chartCat.dataLabels = DataLabelList() 
    chartCat.dataLabels.showPercent = True
    sheet.add_chart(chartCat, "E20")  

   ################################################### 
    sheet.append(['',''])    
    ################################################### 
    
    conteoRLpCatCIS()
    global listaConteoRLpCat
    
    for row in listaConteoRLpCat:
        sheet.append(row)
      
    chartCatRL = BarChart()
    chartCatRL.type = "col"
    chartCatRL.style = 10
    chartCatRL.title = "Categoria por criticidad"
    chartCatRL.y_axis.title = 'Categoria CIS'
    chartCatRL.x_axis.title = '# Findings'
    
    dataCatRL = Reference(sheet, min_col=2, min_row=17, max_row=26, max_col=4)
    labelCatRL = Reference(sheet, min_col=1, min_row=18, max_row=26)
    chartCatRL.add_data(dataCatRL, titles_from_data=True)
    chartCatRL.set_categories(labelCatRL)
    chartCatRL.shape = 4
    
    sh = chartCatRL.series[0]
    sh.graphicalProperties.line.solidFill = "FF0000"
    sh.graphicalProperties.solidFill = "FF0000" 
    
    sm = chartCatRL.series[1]
    sm.graphicalProperties.line.solidFill = "FF8000"
    sm.graphicalProperties.solidFill = "FF8000" 
    
    sl = chartCatRL.series[2]
    sl.graphicalProperties.line.solidFill = "FFFF00"
    sl.graphicalProperties.solidFill = "FFFF00" 
    
    
    chartCatRL.dataLabels = DataLabelList() 
    chartCatRL.dataLabels.showVal = True
    chartCatRL.height = 10 
    chartCatRL.width = 20 
    
    sheet.add_chart(chartCatRL, "E36")
    
    
    
    print('END')
       
    wb.save(cspmPath)    

#########################################CONTEO DE FINDINGS PARA TABLAS#############################################

selectedRisks = []
risks = {}

def createListRisks():
    global selectedRisks  
    selectedRisks = []
    global risks
    for optRL,onOrOff in risks.items():
        if onOrOff.get() == 1:
            selectedRisks.append(optRL)
            print(optRL)
        
def winseleccionRisks():
    critWin = Toplevel(window)    
    critWin.title("Criticidad de las tablas")
    critWin.geometry("300x70")
    menubutton = Menubutton(critWin, text="Criticidades",indicatoron=True, borderwidth=1, relief="raised")    
    menu = Menu(menubutton, tearoff=False)
    menubutton.configure(menu=menu)
    menubutton.grid(column = 0, row = 0,padx=10, pady=10, sticky="nsew")
    
    global risks
    risks = {}
    for risk in ("High", "Medium", "Low"):
        risks[risk] = IntVar(value=0)
        menu.add_checkbutton(label=risk, variable=risks[risk], onvalue=1, offvalue=0, command=createListRisks)
    btn = Button(critWin, text="Generar tablas", command=generarTablas)
    btn.grid(column=1, row=0, padx=10, pady=10, sticky="nsew") 
    
  
def generarTablas():

    wb = openpyxl.load_workbook(cspmPath)
    global subscription
    sheetCSPM = wb[subscription] 
    subIndividual = posibleSubscriptionsParticular.get()
    if subIndividual == 'Todas las suscripciones':
        titleClas = 'Tablas ' + str(subscription) 
        if titleClas not in wb.sheetnames:
            positionS = wb.worksheets.index(wb[subscription]) + 2
            wb.create_sheet(index = positionS, title = titleClas)
            sheetClas = wb[titleClas]
        else: 
            del wb[titleClas] 
           
            positionS = wb.worksheets.index(wb[subscription]) + 2
            wb.create_sheet(index = positionS, title = titleClas)
            sheetClas = wb[titleClas]
    else: 
        titleClas = 'Tablas ' + str(subIndividual)
        if titleClas not in wb.sheetnames:
            positionS = wb.worksheets.index(wb[subscription]) + 2
            wb.create_sheet(index = positionS, title = titleClas)
            sheetClas = wb[titleClas]        
        else: 
            del wb[titleClas] 
           
            positionS = wb.worksheets.index(wb[subscription]) + 2
            wb.create_sheet(index = positionS, title = titleClas)
            sheetClas = wb[titleClas] 
    
    
    referenceCategoryColumn = 1
    referenceSituationColumn = 1
    referenceSubColumn = 1  
    referenceRuleColumn = 1 
    referenceSeverityColumn = 1
    
    global selectedRisks
    if len(selectedRisks) == 0:
        selectedRisks = ['High','Medium','Low']
    
    IAM = 0
    SC = 0
    SA = 0
    DS = 0
    LM = 0
    N = 0
    VM = 0
    OSC = 0
    AS = 0
    countColumn = 1

    tablaIAM = [['Finding','Criticidad','#'],]
    tablaSC = [['Finding','Criticidad','#'],]
    tablaSA = [['Finding','Criticidad','#'],]
    tablaDS = [['Finding','Criticidad','#'],]
    tablaLM = [['Finding','Criticidad','#'],]
    tablaN = [['Finding','Criticidad','#'],]
    tablaVM = [['Finding','Criticidad','#'],]
    tablaOSC = [['Finding','Criticidad','#'],]
    tablaAS = [['Finding','Criticidad','#'],]


    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Category':
            referenceCategoryColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Situation':
            referenceSituationColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Account Name':
            referenceSubColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Rule Title':
            referenceRuleColumn = countColumn
        countColumn += 1

    countColumn = 1    
    while countColumn <= sheetCSPM.max_column:
        if sheetCSPM.cell(row = 1, column = countColumn).value == 'Risk Level':
            referenceSeverityColumn = countColumn
        countColumn += 1

    reglasEvaluadas = []  
    if subIndividual == 'Todas las suscripciones':
        for idxRow in range(2, sheetCSPM.max_row + 1):  
            category = sheetCSPM.cell(row = idxRow, column = referenceCategoryColumn).value
            rule = sheetCSPM.cell(row = idxRow, column = referenceRuleColumn).value
            risk = sheetCSPM.cell(row = idxRow, column = referenceSeverityColumn).value
            situation = sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value

            if situation in ['Pending','New'] and risk in selectedRisks:
                if rule not in reglasEvaluadas:
                    countFinds = 0
                    for comparationIdx in range(2, sheetCSPM.max_row + 1):  
                        if rule == sheetCSPM.cell(row = comparationIdx, column = referenceRuleColumn).value and sheetCSPM.cell(row = comparationIdx, column = referenceSituationColumn).value in  ['Pending','New']:
                            countFinds += 1
                    
                    if category == 'Identity and Access Management':
                        tablaIAM.append([rule,risk,countFinds])
                    elif category == 'Security Center':
                        tablaSC.append([rule,risk,countFinds])
                    elif category == 'Storage Accounts':
                        tablaSA.append([rule,risk,countFinds])
                    elif category == 'Database Services':
                        tablaDS.append([rule,risk,countFinds])
                    elif category == 'Logging and Monitoring':
                        tablaLM.append([rule,risk,countFinds])
                    elif category == 'Networking':
                        tablaN.append([rule,risk,countFinds])   
                    elif category == 'Virtual Machines':
                        tablaVM.append([rule,risk,countFinds])
                    elif category == 'Other Security Considerations':
                        tablaOSC.append([rule,risk,countFinds])
                    else:    
                        tablaAS.append([rule,risk,countFinds])
                reglasEvaluadas.append(rule) 

    else:
        for idxRow in range(2, sheetCSPM.max_row + 1):  
            category = sheetCSPM.cell(row = idxRow, column = referenceCategoryColumn).value
            rule = sheetCSPM.cell(row = idxRow, column = referenceRuleColumn).value
            risk = sheetCSPM.cell(row = idxRow, column = referenceSeverityColumn).value
            situation = sheetCSPM.cell(row = idxRow, column = referenceSituationColumn).value
            subscription = sheetCSPM.cell(row = idxRow, column = referenceSubColumn).value
            
            if situation in ['Pending','New'] and risk in selectedRisks and subscription == subIndividual:
                if rule not in reglasEvaluadas:
                    countFinds = 0
                    for comparationIdx in range(2, sheetCSPM.max_row + 1):  
                        if rule == sheetCSPM.cell(row = comparationIdx, column = referenceRuleColumn).value and sheetCSPM.cell(row = comparationIdx, column = referenceSituationColumn).value in  ['Pending','New'] and sheetCSPM.cell(row = comparationIdx, column = referenceSubColumn).value == subIndividual:
                            countFinds += 1
    
                    if category == 'Identity and Access Management':
                        tablaIAM.append([rule,risk,countFinds])
                    elif category == 'Security Center':
                        tablaSC.append([rule,risk,countFinds])
                    elif category == 'Storage Accounts':
                        tablaSA.append([rule,risk,countFinds])
                    elif category == 'Database Services':
                        tablaDS.append([rule,risk,countFinds])
                    elif category == 'Logging and Monitoring':
                        tablaLM.append([rule,risk,countFinds])
                    elif category == 'Networking':
                        tablaN.append([rule,risk,countFinds])   
                    elif category == 'Virtual Machines':
                        tablaVM.append([rule,risk,countFinds])
                    elif category == 'Other Security Considerations':
                        tablaOSC.append([rule,risk,countFinds])
                    else:    
                        tablaAS.append([rule,risk,countFinds])
    
                reglasEvaluadas.append(rule) 

    sheetClas.append(['Identity and Access Management'])
    for row in tablaIAM:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['Security Center'])
    for row in tablaSC:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['Storage Accounts'])
    for row in tablaSA:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['Database Services'])
    for row in tablaDS:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['Logging and Monitoring'])
    for row in tablaLM:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['Networking'])
    for row in tablaN:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['Virtual Machines'])
    for row in tablaVM:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['Other Security Considerations'])
    for row in tablaOSC:
        sheetClas.append(row)   
     
    sheetClas.append([''])
    
    sheetClas.append(['AppService'])
    for row in tablaAS:
        sheetClas.append(row)   
    
    wb.save(cspmPath)  
    print('Fin recuento')     
  
#########################################################################################################
     
window = Tk() 
   
window.title('CSPM') 
   
window.geometry("700x450") 
   
window.config(background = "white") 

   
label_file_explorer = Label(window, text = "Selecciona el archivo CSPMXX.xlsx y los archivos descargdados del portal (CIS y BM)",width = 100, height = 4,  fg = "blue") 
   
       
search_cspm = Button(window, text = "Continuar CSPM", command = browseFilesCSPM)
new_cspm = Button(window, text = "Nuevo CSPM", command = newCSPMWin)
#label_file_explorer.grid(column = 1, row = 1) 

search_cis = Button(window, text = "Insertar CIS", command = browseFilesCIS)
#label_file_explorer.grid(column = 1, row = 1) 

opcionesFindings = ttk.Combobox(state="readonly",values=['Todas las criticidades','High BM','High CIS','Solo High'])
opcionesFindings.current(0)

addDict = Button(window, text = "Diccionario BM", command = BMDictionary)
#label_file_explorer.grid(column = 1, row = 1) 

search_bm = Button(window, text = "Insertar BM", command = browseFilesBM)
#label_file_explorer.grid(column = 1, row = 1) 

posibleSubscriptions = ttk.Combobox(state="readonly",values=[])
posibleSubscriptionsParticular = ttk.Combobox(state="readonly",values=[])
  
saveSus = Button(window, text='Guardar hoja seleccionada', command=posible_Subscriptions)
#label_file_explorer.grid(column = 1, row = 1) 


evaluacion = Button(window, text = "Evaluacion de findings", command = compareIfPendings)     
#label_file_explorer.grid(column = 1, row = 1) 
setToDone = Button(window, text = "Establecer findings resueltos", command = setToDone)     
generarGraf = Button(window, text = "Generar graficas", command = dibujarGRAFS) 
windowGT = Button(window, text = "Generar tablas", command = winseleccionRisks)     

button_exit = Button(window, text = "Salir", command = window.destroy)     

#button_exit = Button(window, text = "Exit", command = exit)     
#label_file_explorer.grid(column = 1, row = 1) 


label_file_explorer.grid(columnspan = 2, row = 1, sticky="nsew")   

search_cspm.grid(column = 0, row = 2, padx=10, pady=10, sticky="nsew")
new_cspm.grid(row = 2, column = 1, padx=10, pady=10, sticky="nsew") 

search_cis.grid(column = 0, row = 3, padx=10, pady=10, sticky="nsew")
opcionesFindings.grid(column = 1, row = 3, padx=10, pady=10, sticky="nsew")
addDict.grid(column = 1, row = 4, padx=10, pady=10, sticky="nsew") 
search_bm.grid(column = 0, row = 4, padx=10, pady=10, sticky="nsew")
posibleSubscriptions.grid(row = 5, column = 0, padx=10, pady=10, sticky="nsew")
posibleSubscriptionsParticular.grid(row = 5, column = 1, padx=10, pady=10, sticky="nsew")

saveSus.grid(row = 6, column = 0, padx=10, pady=10, sticky="nsew") 

evaluacion.grid(column = 0,row = 7, padx=10, pady=10, sticky="nsew") 
setToDone.grid(column = 1,row = 7, padx=10, pady=10, sticky="nsew") 

generarGraf.grid(column = 0,row = 8, padx=10, pady=10, sticky="nsew") 
windowGT.grid(column = 1,row = 8, padx=10, pady=10, sticky="nsew") 

button_exit.grid(column = 0,row = 9, padx=10, pady=10, sticky="nsew") 

#button_exit.grid(column = 0,row = 9, padx=10, pady=10, sticky="nsew") 

   
window.mainloop() 